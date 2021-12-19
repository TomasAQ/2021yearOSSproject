import openpyxl
import requests
from resource_data.keys import ServiceKey
import urllib
import math
import urllib.request


#global wb , sheet

# 한글을 URL로 넘겨주기 전에 인코딩 진행
def urlEncoding(val):
    encodeResult = urllib.parse.quote(val)
    return encodeResult

def makeXlsx(toulList):
    print(toulList)
    for i in range(len(toulList)):
        if i==0:
            firstXlsx = True
        search(toulList[i] , firstXlsx)
        firstXlsx = False


# 페이지와 검색어를 통해서 검색을 해주는 부분
def search(searKeword , fristXlsx):
    print(searKeword)
    # 한글 인코딩 변경
    encodSearKeword = urlEncoding(searKeword)
    print(encodSearKeword)
    dataJson = "http://api.visitkorea.or.kr/openapi/service/rest/PhotoGalleryService/gallerySearchList?keyword=" + encodSearKeword + "&ServiceKey=" + ServiceKey + "&numOfRows=10&pageNo=" + str(1) + "&MobileOS=ETC&MobileApp=TestApp&_type=json"

    #print(dataJson)
    response = requests.get(dataJson)
    data = response.json()
    #print(data)

    # 검색 관광지 개수
    totalCount = data['response']['body']['totalCount']
    print(totalCount)
    totalpage = math.ceil(totalCount / 10)
    for i in range(1, totalpage + 1):
        print("Page : " + str(i))
        if (i == 1):
            if fristXlsx == True:
                wb , sheet = Exl_head(searKeword)
            else:
                wb , sheet = Exl_head_new(searKeword)
        else:
            dataJson = "http://api.visitkorea.or.kr/openapi/service/rest/PhotoGalleryService/gallerySearchList?keyword=" + encodSearKeword + "&ServiceKey=" + ServiceKey + "&numOfRows=10&pageNo=" + str(i) + "&MobileOS=ETC&MobileApp=TestApp&_type=json"
            response = requests.get(dataJson)
            data = response.json()

        try:
            for j in range(0, 9):
                galTitle = data['response']['body']['items']['item'][j]['galTitle']
                galPhotographyLocation = data['response']['body']['items']['item'][j]['galPhotographyLocation']
                galSearchKeyword = data['response']['body']['items']['item'][j]['galSearchKeyword']
                galViewCount = data['response']['body']['items']['item'][j]['galViewCount']
                galWebImageUrl = data['response']['body']['items']['item'][j]['galWebImageUrl']
                Exl_data(wb, sheet, galTitle, galPhotographyLocation, galSearchKeyword, galViewCount, galWebImageUrl)

        except Exception as e:
            print(e)
            pass




def Exl_head(searKeword):
    global wb
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = searKeword
    sheet['B1'] = "galTitle"
    sheet['C1'] = "galPhotographyLocation"
    sheet['D1'] = "galSearchKeyword"
    sheet['E1'] = "galViewCount"
    sheet['F1'] = "galWebImageUrl"
    return wb, sheet


def Exl_head_new(searKeword):
    wb = openpyxl.load_workbook("관광공사API정리.xlsx")
    sheet = wb.create_sheet(searKeword)
    sheet.title = searKeword
    sheet['B1'] = "galTitle"
    sheet['C1'] = "galPhotographyLocation"
    sheet['D1'] = "galSearchKeyword"
    sheet['E1'] = "galViewCount"
    sheet['F1'] = "galWebImageUrl"
    return wb , sheet


def Exl_data(wb, sheet, galTitle, galPhotographyLocation, galSearchKeyword, galViewCount, galWebImageUrl):
    # todo 이미지 까지 가져오는 방법 생각하기
    #sheet.append(["", galTitle, galPhotographyLocation, galSearchKeyword, galViewCount,  sheet.add_image(urllib.request.urlretrieve(galWebImageUrl))])

    sheet.append(["", galTitle, galPhotographyLocation, galSearchKeyword, galViewCount, galWebImageUrl])
    # 파일 저장
    wb.save("관광공사API정리.xlsx")

if __name__ == '__main_':
    tourList = ["경기도","강원도","충청남도","충청북도","경상북도","경상남도","전라북도","전라남도","서울특별시", "인천광역시", "대전광역시", "대구광역시", "광주광역시", "울산광역시", "부산광역시"]

    tourList2 = ["서울특별시", "인천광역시", "대전광역시", "대구광역시", "광주광역시", "울산광역시", "부산광역시"]
    makeXlsx(tourList)